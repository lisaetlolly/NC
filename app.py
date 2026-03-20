import re
import zipfile
from datetime import date, timedelta
from io import BytesIO
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


STUDIO9_KEYWORD = "STUDIO 9"
PEACH_BAG_KEYWORD = "Peach Tote Bag"


SO_EXCLUDE_KEYWORDS = [
    "包袋",
    "胸针",
    "贺卡",
    "包装服务",
    "手写贺卡",
    "纸袋",
]


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()
    return df


def _first_existing_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    cols = set(map(str, df.columns))
    for c in candidates:
        if c in cols:
            return c
    # 兜底：去空格后匹配
    norm_map = {str(c).strip(): str(c).strip() for c in df.columns}
    for c in candidates:
        if str(c).strip() in norm_map:
            return norm_map[str(c).strip()]
    return None


def _ensure_numeric(df: pd.DataFrame, col: str) -> None:
    if col not in df.columns:
        return
    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)


def _contains_keyword(series: pd.Series, keywords: List[str]) -> pd.Series:
    pattern = "|".join(re.escape(k) for k in keywords if k)
    if not pattern:
        return pd.Series([False] * len(series), index=series.index)
    return series.astype(str).str.contains(pattern, case=False, na=False, regex=True)


def _violent_clean_key_series(series: pd.Series) -> pd.Series:
    """
    最暴力的 key 清理：strip + 去掉不可见空白 + 去掉 Excel 浮点尾部 .0。
    """
    s = series.astype(str)
    s = s.replace("\u00a0", " ").replace("\u3000", " ")
    s = s.str.strip()
    # 去掉尾部 .0（以及 .0...）
    s = s.str.replace(r"(\d+)\.0+$", r"\1", regex=True)
    # 再次去掉空格
    s = s.str.strip()
    # 统一空值
    s = s.replace({"nan": "", "NaN": "", "None": ""})
    return s


def _clean_th_key(external_single_no: object) -> str:
    s = "" if external_single_no is None else str(external_single_no)
    # 去除不可见空白 + 去掉可能的 .0 尾巴（Excel 常见）
    s = s.replace("\u00a0", " ").replace("\u3000", " ").strip()
    s = re.sub(r"(\d+)\.0+$", r"\1", s)
    s = s.strip()
    if not s:
        return ""
    # TH开头且后接至少11位数字：TH + 11 digits
    if s.startswith("TH"):
        digits = re.sub(r"\D+", "", s[2:])
        # 取“后续的11位数字”（更宽松：取最后11位，避免前缀/中间补位导致错位）
        return digits[-11:] if len(digits) >= 11 else digits
    # 其他情况：抓取首次11位连续数字
    m = re.search(r"(\d{11})", s)
    if m:
        return m.group(1)
    digits = re.sub(r"\D+", "", s)
    return digits[-11:] if len(digits) >= 11 else digits


def _normalize_merge_key(x: object) -> str:
    """
    为 merge 做归一化：
    1) 转成字符串并 strip
    2) 提取所有数字并拼接
    3) 去除前导 0（全是 0 时保留一个 0）
    """
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\u00a0", " ").replace("\u3000", " ").strip()
    # Excel 常把数字写成 12345.0，直接把尾部 .0 去掉，避免误拼出多一位 0
    s = re.sub(r"(\d+)\.0+$", r"\1", s)
    if s in {"nan", "NaN", "None"}:
        return ""
    s = re.sub(r"\s+", "", s)
    digits = "".join(re.findall(r"\d+", s))
    if not digits:
        return s
    digits = digits.lstrip("0")
    return digits if digits != "" else "0"


def _parse_first_order_id(x: object) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    s = str(x).strip()
    if not s:
        return ""
    # 可能存在多个订单号：逗号/分号/中文逗号/空格/换行等分隔
    tokens = re.split(r"[,\s;，/|]+", s)
    tokens = [t.strip() for t in tokens if t.strip()]
    return tokens[0] if tokens else s


def _shop_bucket(shop: str) -> str:
    s = "" if shop is None or (isinstance(shop, float) and np.isnan(shop)) else str(shop).strip()
    if not s or s == "N/A" or s == "N／A":
        return "HAY-Tmall"
    if any(k in s for k in ["小红书", "HAY旗舰店", "RED"]):
        return "RED"
    if STUDIO9_KEYWORD in s:
        return "JD-STUDIO 9"
    if "天猫" in s or "Tmall" in s or "京东" in s:
        return "HAY-Tmall"
    # 兜底：按你的要求默认天猫/HAY-Tmall
    return "HAY-Tmall"


def _build_aux_map(aux_df: Optional[pd.DataFrame]) -> Dict[str, Tuple[str, str]]:
    """
    返回：完整编码 -> (主计量单位, 辅计量单位)
    严禁截断前6位，必须以 Full Code/完整编码精确匹配。
    """
    if aux_df is None or aux_df.empty:
        return {}

    aux_df = _normalize_columns(aux_df)

    code_col = _first_existing_col(
        aux_df,
        ["Full Code", "匹配物料编码", "物料编码", "编码", "物料代码", "电商系统物料编码"],
    )
    main_unit_col = _first_existing_col(aux_df, ["Unit", "主计量单位", "主单位", "主计量"])
    sub_unit_col = _first_existing_col(aux_df, ["辅计量单位", "辅单位"])

    if not main_unit_col:
        main_unit_col = _first_existing_col(aux_df, ["计量单位", "单位"])
    if not sub_unit_col:
        sub_unit_col = main_unit_col

    if not code_col or not main_unit_col:
        return {}

    result: Dict[str, Tuple[str, str]] = {}
    for _, r in aux_df.iterrows():
        c = r.get(code_col, "")
        if pd.isna(c):
            continue
        if isinstance(c, float) and float(c).is_integer():
            c_str = str(int(c))
        else:
            c_str = str(c).strip()
        c_str = re.sub(r"\.0+$", "", c_str)
        if not c_str:
            continue

        main_u = str(r.get(main_unit_col, "PCS")).strip()
        sub_u = str(r.get(sub_unit_col, main_u)).strip()
        if not main_u or main_u.lower() == "nan":
            main_u = "PCS"
        if not sub_u or sub_u.lower() == "nan":
            sub_u = main_u

        # 用完整编码作为 key
        result[c_str] = (main_u, sub_u)

    return result


def _is_int_like(x: object) -> bool:
    try:
        if isinstance(x, (int, np.integer)):
            return True
        if isinstance(x, float) and float(x).is_integer():
            return True
    except Exception:
        return False
    return False


def _enrich_so_df(
    so1_df: pd.DataFrame,
    so2_df: pd.DataFrame,
    aux_map: Dict[str, Tuple[str, str]],
    debug: bool = False,
) -> pd.DataFrame:
    so1_df = _normalize_columns(so1_df)
    so2_df = _normalize_columns(so2_df)

    so1_df = so1_df.copy()
    so2_df = so2_df.copy()

    # 依据你的要求：双主键防止笛卡尔积膨胀
    so2_ext_col = _first_existing_col(so2_df, ["外部单号", "外部订单号", "外部单号2", "单号"])
    so2_sku_col = _first_existing_col(so2_df, ["货品", "商品编码", "SKU编码"])
    so1_out_col = _first_existing_col(so1_df, ["出仓单号", "出仓单", "出库单号"])
    so1_sku_code_col = _first_existing_col(so1_df, ["商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
    oms_col = _first_existing_col(so2_df, ["OMS"])

    if not so2_ext_col or not so2_sku_col or not so1_out_col or not so1_sku_code_col or not oms_col:
        if debug:
            st.warning(
                f"SO无法定位双键：so2_ext={so2_ext_col}, so2_sku={so2_sku_col}, so1_out={so1_out_col}, so1_sku={so1_sku_code_col}, oms_col={oms_col}"
            )
        return pd.DataFrame()

    # 强制清理四个 join key + OMS
    so1_df[so1_out_col] = _violent_clean_key_series(so1_df[so1_out_col])
    so1_df[so1_sku_code_col] = _violent_clean_key_series(so1_df[so1_sku_code_col])
    so2_df[so2_ext_col] = _violent_clean_key_series(so2_df[so2_ext_col])
    so2_df[so2_sku_col] = _violent_clean_key_series(so2_df[so2_sku_col])
    so2_df[oms_col] = _violent_clean_key_series(so2_df[oms_col])

    # 仅聚水潭发货明细
    so2_filt = so2_df.loc[so2_df[oms_col].astype(str).str.contains("聚水潭", na=False)].copy()
    if debug:
        st.write(f"👉 识别到底表1(聚水潭出库)，总行数: {len(so1_df)}")
        st.write(f"👉 识别到底表2(WMS发货)，总行数: {len(so2_df)}")
        st.write(f"👉 底表2 过滤聚水潭后，剩余行数: {len(so2_filt)}")
    if so2_filt.empty:
        return pd.DataFrame()

    merged = pd.merge(
        so2_filt,
        so1_df,
        how="left",
        left_on=[so2_ext_col, so2_sku_col],
        right_on=[so1_out_col, so1_sku_code_col],
        suffixes=("_底2", "_底1"),
    )

    # 去重防卫：避免 so1/so2 侧存在重复键导致行重复
    merged = merged.drop_duplicates(subset=[so2_ext_col, so2_sku_col, so1_out_col, so1_sku_code_col])

    if debug:
        st.write(f"👉 两表 Merge 匹配后（双键防膨胀），成功合并的行数: {len(merged)}")

    # 必备字段猜测
    # RT：强力候选数量/金额字段（不同模板可能叫法不同）
    qty_col = _first_existing_col(
        merged,
        [
            "实发数量",
            "数量",
            "退货数量",
            "售后数量",
            "退款数量",
            "收货数量",
            "入库数量",
        ],
    )
    amount_col = _first_existing_col(
        merged,
        [
            "实发金额",
            "金额",
            "退货金额",
            "售后金额",
            "退款金额",
            "退款总金额",
            "实退金额",
            "含税金额",
            "价税合计",
        ],
    )
    ship_income_col = _first_existing_col(merged, ["运费收入分摊", "运费收入分摊金额", "运费分摊"])
    ship_fee_col = _first_existing_col(merged, ["运费金额", "运费"])
    sku_name_col = _first_existing_col(merged, ["商品简称", "商品名称", "商品描述"])
    sku_code_col = _first_existing_col(merged, ["商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
    shop_col = _first_existing_col(merged, ["店铺", "店铺名称", "平台店铺"])
    order_col = so1_out_col
    online_order_col = _first_existing_col(merged, ["线上订单号", "客户订单号", "线上订单", "订单号线上"])

    if not qty_col or not amount_col:
        if debug:
            st.warning(f"SO无法定位金额/数量列：qty_col={qty_col}, amount_col={amount_col}")
        return pd.DataFrame()

    # 剔除非保仓：实发数量==0 的数据剔除
    _ensure_numeric(merged, qty_col)
    merged = merged.loc[merged[qty_col] != 0].copy()
    if debug:
        st.write(f"SO：实发数量!=0后 rows={len(merged)}")
    if merged.empty:
        return pd.DataFrame()

    if ship_income_col and amount_col:
        _ensure_numeric(merged, amount_col)
        _ensure_numeric(merged, ship_income_col)
        merged["实际支付金额"] = pd.to_numeric(merged[amount_col], errors="coerce").fillna(0) - pd.to_numeric(merged[ship_income_col], errors="coerce").fillna(0)
    else:
        # 缺少运费分摊时，直接等于实发金额
        merged["实际支付金额"] = pd.to_numeric(merged[amount_col], errors="coerce").fillna(0)
        if not ship_income_col:
            merged["运费收入分摊"] = 0

    if not ship_income_col:
        ship_income_col = "运费收入分摊"
        merged[ship_income_col] = 0
    else:
        merged[ship_income_col] = pd.to_numeric(merged[ship_income_col], errors="coerce").fillna(0)

    # 排除商品
    if sku_name_col:
        mask_exclude = _contains_keyword(merged[sku_name_col], SO_EXCLUDE_KEYWORDS)
        merged = merged.loc[~mask_exclude].copy()
        if debug:
            st.write(f"SO：剔除商品关键词后 rows={len(merged)}")
    if merged.empty:
        return pd.DataFrame()

    # 京东特殊规则：STUDIO 9 且单笔总金额 >= 799，Peach Tote Bag 强制设为0
    if shop_col and sku_name_col:
        merged[shop_col] = merged[shop_col].astype(str)
        merged["__是否STUDIO9__"] = merged[shop_col].str.contains(STUDIO9_KEYWORD, na=False)
        merged["__是否PeachBag__"] = merged[sku_name_col].astype(str).str.contains(PEACH_BAG_KEYWORD, na=False)

        order_total = merged.groupby(order_col)["实际支付金额"].sum()
        eligible_orders = set(order_total.loc[order_total >= 799].index.tolist())

        adj_mask = merged["__是否STUDIO9__"] & merged["__是否PeachBag__"] & merged[order_col].isin(eligible_orders)
        if adj_mask.any():
            # 将该包袋金额及运费分摊置0，避免出现实际支付金额为负
            merged.loc[adj_mask, amount_col] = 0
            if ship_income_col and ship_income_col in merged.columns:
                merged.loc[adj_mask, ship_income_col] = 0
            elif "运费收入分摊" in merged.columns:
                merged.loc[adj_mask, "运费收入分摊"] = 0
            merged.loc[adj_mask, "实际支付金额"] = 0
        merged.drop(columns=["__是否STUDIO9__", "__是否PeachBag__"], errors="ignore", inplace=True)

    # 输出统一列（运费行由报表层 H101 生成，这里只保留运费金额字段）
    if not sku_code_col:
        merged["商品编码"] = ""
        sku_code_col = "商品编码"
    else:
        merged.rename(columns={sku_code_col: "商品编码"}, inplace=True)

    merged.rename(columns={qty_col: "实发数量"}, inplace=True)
    merged.rename(columns={order_col: "订单号"}, inplace=True)
    if shop_col and shop_col in merged.columns:
        merged["店铺"] = merged[shop_col]
    else:
        merged["店铺"] = "N/A"

    if online_order_col and online_order_col in merged.columns:
        merged["线上订单号"] = merged[online_order_col].apply(_parse_first_order_id)
    else:
        merged["线上订单号"] = merged["订单号"].apply(_parse_first_order_id)

    # 运费收入分摊列确保存在
    if ship_income_col not in merged.columns:
        merged["运费收入分摊"] = 0

    # 不强依赖：后续报表只用“商品编码/实发数量/实际支付金额/运费金额/线上订单号/店铺/订单号”
    if ship_fee_col and ship_fee_col in merged.columns:
        merged["运费金额"] = pd.to_numeric(merged[ship_fee_col], errors="coerce").fillna(0)
    else:
        merged["运费金额"] = 0.0

    return merged[["店铺", "商品编码", "实发数量", "实际支付金额", "运费金额", "线上订单号", "订单号"]].copy()


def _enrich_rt_df(
    rt3_df: pd.DataFrame,
    rt4_df: pd.DataFrame,
    aux_map: Dict[str, Tuple[str, str]],
    debug: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    返回：(rt_main_df, rt_nonbao_df)
    rt_main_df: 实发数量 != 0
    rt_nonbao_df: 实发数量 == 0
    """
    rt3_df = _normalize_columns(rt3_df)
    rt4_df = _normalize_columns(rt4_df)

    oms_col = _first_existing_col(rt4_df, ["OMS"])
    external_key = _first_existing_col(rt4_df, ["外部单号", "外部订单号", "单号"])
    after_key = _first_existing_col(rt3_df, ["售后单号", "售后订单号", "单号"])

    if not oms_col or not external_key or not after_key:
        if debug:
            st.warning(f"RT无法定位合并键：oms_col={oms_col}, external_key={external_key}, after_key={after_key}")
        return pd.DataFrame(), pd.DataFrame()

    # 同样在 merge 前对 join key 做强转+清理
    rt3_df[after_key] = _violent_clean_key_series(rt3_df[after_key])
    rt4_df[external_key] = _violent_clean_key_series(rt4_df[external_key])
    rt4_df[oms_col] = _violent_clean_key_series(rt4_df[oms_col])

    # 仅聚水潭收货明细（宽容：contains）
    rt4_filt = rt4_df.loc[rt4_df[oms_col].astype(str).str.contains("聚水潭", na=False)].copy()
    if debug:
        st.write(f"👉 识别到底表3(聚水潭退货)，总行数: {len(rt3_df)}")
        st.write(f"👉 识别到底表4(WMS收货)，总行数: {len(rt4_df)}")
        st.write(f"👉 底表4 过滤聚水潭后，剩余行数: {len(rt4_filt)}")
    if rt4_filt.empty:
        return pd.DataFrame(), pd.DataFrame()

    rt4_filt["__THKey__"] = rt4_filt[external_key].apply(_clean_th_key).astype(str).str.strip()
    rt4_filt = rt4_filt.loc[rt4_filt["__THKey__"] != ""].copy()
    if debug:
        st.write(f"RT：THKey!=空后 rows={len(rt4_filt)}")
    if rt4_filt.empty:
        return pd.DataFrame(), pd.DataFrame()

    merged = rt3_df.merge(
        rt4_filt,
        left_on=after_key,
        right_on="__THKey__",
        how="inner",
        suffixes=("_底3", "_底4"),
    )
    if debug:
        st.write(f"👉 两表 Merge 匹配后，成功合并的行数: {len(merged)}")
    if merged.empty:
        if debug:
            st.write("RT：merge后为空（inner join 找不到匹配TH键/售后单号）")
        # 宽松匹配：用“提取数字+去前导0”的归一化键再 merge 一次
        rt3_tmp = rt3_df.copy()
        rt4_tmp = rt4_filt.copy()
        rt3_tmp["__RTAfKeyNorm__"] = rt3_tmp[after_key].map(_normalize_merge_key)
        rt4_tmp["__RTTHKeyNorm__"] = rt4_tmp["__THKey__"].map(_normalize_merge_key)

        rt3_tmp = rt3_tmp.loc[rt3_tmp["__RTAfKeyNorm__"] != ""].copy()
        rt4_tmp = rt4_tmp.loc[rt4_tmp["__RTTHKeyNorm__"] != ""].copy()

        if debug:
            inter = set(rt3_tmp["__RTAfKeyNorm__"].unique()) & set(rt4_tmp["__RTTHKeyNorm__"].unique())
            st.write(f"RT：归一化键交集数量={len(inter)}")

        merged2 = rt3_tmp.merge(
            rt4_tmp,
            left_on="__RTAfKeyNorm__",
            right_on="__RTTHKeyNorm__",
            how="inner",
            suffixes=("_底3", "_底4"),
        )
        if merged2.empty:
            if debug:
                st.write("RT：归一化 merge 仍为空")
            return pd.DataFrame(), pd.DataFrame()
        merged = merged2

    qty_col = _first_existing_col(merged, ["实发数量", "数量"])
    amount_col = _first_existing_col(merged, ["实发金额", "金额"])
    ship_income_col = _first_existing_col(merged, ["运费收入分摊", "运费收入分摊金额", "运费分摊"])
    ship_fee_col = _first_existing_col(merged, ["运费金额", "运费"])
    sku_name_col = _first_existing_col(merged, ["商品简称", "商品名称", "商品描述"])
    sku_code_col = _first_existing_col(merged, ["商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
    shop_col = _first_existing_col(merged, ["店铺", "店铺名称", "平台店铺"])
    online_order_col = _first_existing_col(merged, ["线上订单号", "客户订单号", "线上订单", "订单号线上", "线上订单编号"])

    if debug:
        st.write(f"RT字段识别：qty_col={qty_col}, amount_col={amount_col}, ship_income_col={ship_income_col}")

    if not qty_col or not amount_col:
        if debug:
            qty_candidates = [c for c in merged.columns if "数量" in str(c)]
            amt_candidates = [c for c in merged.columns if ("金额" in str(c) or "价税" in str(c) or "退款" in str(c))]
            st.warning(
                f"RT无法定位金额/数量列：qty_col={qty_col}, amount_col={amount_col}"
            )
            st.write(f"RT merged 含有数量相关的列名（前20个）: {qty_candidates[:20]}")
            st.write(f"RT merged 含有金额相关的列名（前20个）: {amt_candidates[:20]}")
        return pd.DataFrame(), pd.DataFrame()

    _ensure_numeric(merged, qty_col)
    _ensure_numeric(merged, amount_col)

    # 实际支付金额
    if ship_income_col:
        _ensure_numeric(merged, ship_income_col)
        merged["实际支付金额"] = pd.to_numeric(merged[amount_col], errors="coerce").fillna(0) - pd.to_numeric(merged[ship_income_col], errors="coerce").fillna(0)
    else:
        merged["实际支付金额"] = pd.to_numeric(merged[amount_col], errors="coerce").fillna(0)
        merged["运费收入分摊"] = 0
        ship_income_col = "运费收入分摊"

    # 商品排除
    if sku_name_col:
        mask_exclude = _contains_keyword(merged[sku_name_col], SO_EXCLUDE_KEYWORDS)
        merged = merged.loc[~mask_exclude].copy()
        if debug:
            st.write(f"RT：剔除商品关键词后 rows={len(merged)}")
    if merged.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 京东特殊规则同SO（STUDIO 9 & 单笔>=799 & Peach Tote Bag置0）
    if shop_col and sku_name_col:
        merged[shop_col] = merged[shop_col].astype(str)
        merged["__是否STUDIO9__"] = merged[shop_col].str.contains(STUDIO9_KEYWORD, na=False)
        merged["__是否PeachBag__"] = merged[sku_name_col].astype(str).str.contains(PEACH_BAG_KEYWORD, na=False)

        order_col = after_key
        order_total = merged.groupby(order_col)["实际支付金额"].sum()
        eligible_orders = set(order_total.loc[order_total >= 799].index.tolist())
        adj_mask = merged["__是否STUDIO9__"] & merged["__是否PeachBag__"] & merged[order_col].isin(eligible_orders)
        if adj_mask.any():
            merged.loc[adj_mask, amount_col] = 0
            if ship_income_col and ship_income_col in merged.columns:
                merged.loc[adj_mask, ship_income_col] = 0
            elif "运费收入分摊" in merged.columns:
                merged.loc[adj_mask, "运费收入分摊"] = 0
            merged.loc[adj_mask, "实际支付金额"] = 0
        merged.drop(columns=["__是否STUDIO9__", "__是否PeachBag__"], errors="ignore", inplace=True)

    # 运费H101切分
    if ship_fee_col:
        _ensure_numeric(merged, ship_fee_col)
        fee_mask = pd.to_numeric(merged[ship_fee_col], errors="coerce").fillna(0) > 0
        if fee_mask.any():
            h101_rows = merged.loc[fee_mask].copy()
            h101_rows[qty_col] = 1
            if ship_income_col in h101_rows.columns:
                h101_rows[ship_income_col] = 0
            h101_rows["实际支付金额"] = pd.to_numeric(h101_rows[ship_fee_col], errors="coerce").fillna(0)
            # 商品编码切分：写入到“原始商品编码列”，避免后续 rename 覆盖 H101 设置
            if sku_code_col and sku_code_col in h101_rows.columns:
                h101_rows[sku_code_col] = "H101"
            else:
                h101_rows["商品编码"] = "H101"
            merged = pd.concat([merged, h101_rows], ignore_index=True)

    if not sku_code_col:
        merged["商品编码"] = ""
        sku_code_col = "商品编码"
    else:
        merged.rename(columns={sku_code_col: "商品编码"}, inplace=True)

    merged.rename(columns={qty_col: "实发数量"}, inplace=True)
    merged.rename(columns={after_key: "订单号"}, inplace=True)
    if shop_col and shop_col in merged.columns:
        merged["店铺"] = merged[shop_col]
    else:
        merged["店铺"] = "N/A"

    if online_order_col and online_order_col in merged.columns:
        merged["线上订单号"] = merged[online_order_col].apply(_parse_first_order_id)
    else:
        merged["线上订单号"] = merged["订单号"].apply(_parse_first_order_id)

    if "实际支付金额" not in merged.columns:
        merged["实际支付金额"] = 0

    # 输出统一列
    base = merged[["店铺", "商品编码", "实发数量", "实际支付金额", "线上订单号", "订单号"]].copy()
    nonbao_df = base.loc[base["实发数量"] == 0].copy()
    main_df = base.loc[base["实发数量"] != 0].copy()
    return main_df, nonbao_df


def _enrich_rt_df_strict(
    rt3_df: pd.DataFrame,
    rt4_df: pd.DataFrame,
    debug: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    RT严格模式（按你的文档字段标准，不再对“金额/数量列”做兜底猜测）：
    - 金额：来自底表3 `退货金额`
    - 数量：来自底表4 `收货数量`
    - 匹配键：底表4 `外部单号` 去掉 TH 后取前11位数字 -> 底表3 `售后单号`
    """
    # 额外兜底：再次清理表头空格（你在 main 已经做了立即清理）
    rt3_df.columns = rt3_df.columns.astype(str).str.strip()
    rt4_df.columns = rt4_df.columns.astype(str).str.strip()

    after_key_col = "售后单号"  # 聚水潭
    rt3_amount_col = "退货金额"  # 你指出：底表3 金额列实际叫该名字
    rt4_external_col = "外部单号"  # WMS
    rt4_qty_col = "收货数量"  # WMS

    # 严格放行标准：只要底表3有 退货金额、底表4有 收货数量，就允许继续计算
    missing_core = []
    for c in [rt3_amount_col]:
        if c not in rt3_df.columns:
            missing_core.append(f"底表3缺列:{c}")
    for c in [rt4_qty_col]:
        if c not in rt4_df.columns:
            missing_core.append(f"底表4缺列:{c}")
    if missing_core:
        if debug:
            st.warning("RT严格模式：必需字段缺失：" + "；".join(missing_core))
            st.write(f"底表3列名样本：{list(rt3_df.columns[:40])}")
            st.write(f"底表4列名样本：{list(rt4_df.columns[:40])}")
        return pd.DataFrame(), pd.DataFrame()

    oms_col = _first_existing_col(rt4_df, ["OMS"])
    if not oms_col:
        # 没有 OMS 也只能全量当作聚水潭
        oms_col = None

    if debug:
        st.write(f"👉 识别到底表3(聚水潭退货)，总行数: {len(rt3_df)}")
        st.write(f"👉 识别到底表4(WMS收货)，总行数: {len(rt4_df)}")
        st.write(
            "👉 RT严格模式字段存在性："
            + f"rt3[{rt3_amount_col}]={'是' if rt3_amount_col in rt3_df.columns else '否'}；"
            + f"rt4[{rt4_qty_col}]={'是' if rt4_qty_col in rt4_df.columns else '否'}；"
            + f"rt3[{after_key_col}]={'是' if after_key_col in rt3_df.columns else '否'}；"
            + f"rt4[{rt4_external_col}]={'是' if rt4_external_col in rt4_df.columns else '否'}；"
            + f"rt3[商品编码]={'是' if '商品编码' in rt3_df.columns else '否'}"
        )

    if oms_col:
        rt4_filt = rt4_df.loc[rt4_df[oms_col].astype(str).str.contains("聚水潭", na=False)].copy()
    else:
        rt4_filt = rt4_df.copy()
    if rt4_filt.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 如果关键匹配列缺失：不做 join，至少保留 WMS 行（金额无法对齐则置 0）
    if after_key_col not in rt3_df.columns or rt4_external_col not in rt4_filt.columns:
        if debug:
            st.warning(
                f"RT严格模式：缺少匹配键，跳过join；missing after_key_col={after_key_col in rt3_df.columns}, "
                f"rt4_external_col={rt4_external_col in rt4_filt.columns}"
            )
        sku_code_col = _first_existing_col(rt4_df, ["货品", "商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
        base_out = pd.DataFrame(
            {
                "店铺": "天猫",
                "商品编码": rt4_filt[sku_code_col] if sku_code_col and sku_code_col in rt4_filt.columns else "",
                "实发数量": -pd.to_numeric(rt4_filt[rt4_qty_col], errors="coerce").fillna(0.0),
                "实际支付金额": 0.0,
                "线上订单号": "",
                "订单号": "",
            }
        )
        nonbao_df = base_out.loc[base_out["实发数量"] == 0].copy()
        main_df = base_out.loc[base_out["实发数量"] != 0].copy()
        return main_df, nonbao_df

    # 按要求：匹配前统一类型：astype(str).str.strip().str.replace(r'\\.0$', '', regex=True)
    rt3_df["__after_norm__"] = (
        rt3_df[after_key_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    )
    rt4_filt["__ext_norm__"] = (
        rt4_filt[rt4_external_col].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    )

    # TH 单号清洗：若以 TH 开头，去掉 TH；然后仅取前 11 位数字
    # （对非 TH 的情况：同样提取数字并截取前 11 位，用于尽量匹配售后单号）
    no_th = rt4_filt["__ext_norm__"].str.replace(r"(?i)^th", "", regex=True)
    digits_only = no_th.str.replace(r"\D+", "", regex=True)
    rt4_filt["__match_no__"] = digits_only.str.slice(0, 11)

    if debug:
        st.write(f"👉 底表4 过滤聚水潭后，剩余行数: {len(rt4_filt)}")
        st.write(f"👉 __match_no__!=空后 rows={(rt4_filt['__match_no__']!='').sum()}")

    rt4_filt["__RT4_ID__"] = range(len(rt4_filt))

    # 双键防膨胀：
    # WMS(底表4)匹配键：['匹配单号','货品']
    # 聚水潭(底表3)匹配键：['售后单号','商品编码']（如没有商品编码，则退化为单键+去重防卫）
    wms_item_key = _first_existing_col(rt4_filt, ["货品"])
    rt3_item_key = "商品编码" if "商品编码" in rt3_df.columns else None

    merge_left_keys = ["__match_no__"]
    merge_right_keys = ["__after_norm__"]
    if wms_item_key and rt3_item_key:
        merge_left_keys.append(wms_item_key)
        merge_right_keys.append(rt3_item_key)

    rt_merged = pd.merge(
        rt4_filt,
        rt3_df,
        how="left",
        left_on=merge_left_keys,
        right_on=merge_right_keys,
        suffixes=("_wms", "_rt3"),
    )

    # 去重防卫：防止底表3在同售后单号存在多行导致笛卡尔膨胀
    dedupe_subset = ["__RT4_ID__"]
    if wms_item_key:
        dedupe_subset.append(wms_item_key)
    rt_merged = rt_merged.drop_duplicates(subset=dedupe_subset, keep="first")

    if debug:
        st.write(f"👉 两表 Merge 匹配后，成功合并的行数: {len(rt_merged)}")

    # 固定列硬抓：仅按你给的标准字段计算并转负数
    # 价税合计 = rt_merged['退货金额'] * -1
    # 数量 = rt_merged['收货数量'] * -1
    rt_merged["__amt__"] = pd.to_numeric(rt_merged[rt3_amount_col], errors="coerce").fillna(0.0) * -1.0
    rt_merged["__qty__"] = pd.to_numeric(rt_merged[rt4_qty_col], errors="coerce").fillna(0.0) * -1.0

    rt_merged["实发数量"] = rt_merged["__qty__"]
    rt_merged["实际支付金额"] = rt_merged["__amt__"]

    # 其它列用于报表输出（不影响金额/数量的严格标准）
    sku_code_col = _first_existing_col(rt4_df, ["货品", "商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
    shop_col = _first_existing_col(rt3_df, ["店铺", "店铺名称", "平台店铺"])
    online_order_col = _first_existing_col(rt3_df, ["线上订单号", "线上订单", "客户订单号", "订单号线上", "线上订单编号"])

    base_out = pd.DataFrame(
        {
            "店铺": rt_merged[shop_col] if shop_col and shop_col in rt_merged.columns else np.nan,
            "商品编码": rt_merged[sku_code_col] if sku_code_col and sku_code_col in rt_merged.columns else "",
            "实发数量": rt_merged["实发数量"],
            "实际支付金额": rt_merged["实际支付金额"],
            "线上订单号": rt_merged[online_order_col] if online_order_col and online_order_col in rt_merged.columns else "",
            "订单号": rt_merged.get(after_key_col, ""),
        }
    )
    # 店铺 N/A 兜底：统一填充为天猫，归入天猫报表桶
    base_out["店铺"] = base_out["店铺"].fillna("天猫")
    base_out["店铺"] = base_out["店铺"].replace({"N/A": "天猫", "N／A": "天猫", "": "天猫"})

    nonbao_df = base_out.loc[base_out["实发数量"] == 0].copy()
    main_df = base_out.loc[base_out["实发数量"] != 0].copy()

    # 输出前补齐 NaN：字符串为空，数值为0（避免 Excel 写入 NaN）
    for col in ["店铺", "商品编码", "线上订单号", "订单号"]:
        if col in main_df.columns:
            main_df[col] = main_df[col].fillna("").astype(str)
        if col in nonbao_df.columns:
            nonbao_df[col] = nonbao_df[col].fillna("").astype(str)

    for col in ["实发数量", "实际支付金额"]:
        if col in main_df.columns:
            main_df[col] = pd.to_numeric(main_df[col], errors="coerce").fillna(0.0)
        if col in nonbao_df.columns:
            nonbao_df[col] = pd.to_numeric(nonbao_df[col], errors="coerce").fillna(0.0)

    return main_df, nonbao_df


def _enrich_rt_df_v2(
    rt3_df: pd.DataFrame,
    rt4_df: pd.DataFrame,
    aux_map: Dict[str, Tuple[str, str]],
    debug: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    RT 强化版：
    - 用 WMS TH 清洗后的外部单号 left join 聚水潭售后单号
    - 对未匹配行，再用 WMS 外部单号匹配聚水潭 线上订单号/内部订单号
    - 金额/数量字段按要求优先级取值，并在输出阶段全部转为负数
    """
    return _enrich_rt_df_strict(rt3_df, rt4_df, debug=debug)

    oms_col = _first_existing_col(rt4_df, ["OMS"])
    wms_external_col = _first_existing_col(rt4_df, ["外部单号", "外部订单号", "外部单号2", "单号"])
    rt3_sa_after_col = _first_existing_col(rt3_df, ["售后单号", "售后订单号", "单号"])

    if not oms_col or not wms_external_col or not rt3_sa_after_col:
        if debug:
            st.warning(f"RT_v2无法定位合并键：oms={oms_col}, wms_ext={wms_external_col}, rt3_after={rt3_sa_after_col}")
        return pd.DataFrame(), pd.DataFrame()

    # 取数量/金额字段（按你的优先级）
    wms_qty_col = _first_existing_col(rt4_df, ["收货数量", "退货数量", "数量"])
    wms_amount_col = _first_existing_col(rt4_df, ["退货金额", "实际入库金额", "线上申请金额", "金额", "退款金额", "实退金额"])
    wms_ship_fee_col = _first_existing_col(rt4_df, ["运费金额", "运费"])

    # 商品编码优先用 WMS 的货品列
    wms_sku_col = _first_existing_col(rt4_df, ["货品", "商品编码", "物料编码", "SKU编码", "电商系统物料编码"])

    # 退货表排除关键词用尽可能接近的列
    wms_sku_name_col = _first_existing_col(rt4_df, ["货品名称", "商品名称", "商品简称", "商品描述", "货品"])

    # 聚水潭侧用于输出店铺/线上订单号/内部订单号
    rt3_shop_col = _first_existing_col(rt3_df, ["店铺", "店铺名称", "平台店铺"])
    rt3_online_col = _first_existing_col(rt3_df, ["线上订单号", "线上订单", "客户订单号", "订单号线上", "线上订单编号"])
    rt3_internal_col = _first_existing_col(rt3_df, ["内部订单号", "内部订单", "内部订单编号"])

    missing_required = []
    if not wms_qty_col:
        missing_required.append("收货数量/退货数量")
    if not wms_amount_col:
        missing_required.append("退货金额/实际入库金额/线上申请金额")
    if not wms_sku_col:
        missing_required.append("货品/商品编码")
    if missing_required:
        if debug:
            st.warning(f"RT_v2缺少必要字段：{missing_required}")
        return pd.DataFrame(), pd.DataFrame()

    # 强制清理关键列
    rt4_df[wms_external_col] = _violent_clean_key_series(rt4_df[wms_external_col])
    rt4_df[oms_col] = _violent_clean_key_series(rt4_df[oms_col])
    rt3_df[rt3_sa_after_col] = _violent_clean_key_series(rt3_df[rt3_sa_after_col])
    if rt3_online_col:
        rt3_df[rt3_online_col] = _violent_clean_key_series(rt3_df[rt3_online_col])
    if rt3_internal_col:
        rt3_df[rt3_internal_col] = _violent_clean_key_series(rt3_df[rt3_internal_col])
    if rt3_shop_col:
        rt3_df[rt3_shop_col] = _violent_clean_key_series(rt3_df[rt3_shop_col])

    # OMS过滤
    rt4_filt = rt4_df.loc[rt4_df[oms_col].astype(str).str.contains("聚水潭", na=False)].copy()
    if rt4_filt.empty:
        return pd.DataFrame(), pd.DataFrame()

    rt4_filt["__WMSExternalNorm__"] = rt4_filt[wms_external_col]
    rt4_filt["__WMSKeyTH__"] = rt4_filt[wms_external_col].apply(_clean_th_key)
    rt4_filt["__WMSKeyTH__"] = _violent_clean_key_series(rt4_filt["__WMSKeyTH__"])
    if debug:
        st.write(f"👉 识别到底表3(聚水潭退货)，总行数: {len(rt3_df)}")
        st.write(f"👉 识别到底表4(WMS收货)，总行数: {len(rt4_df)}")
        st.write(f"👉 底表4 过滤聚水潭后，剩余行数: {len(rt4_filt)}")
        st.write(f"👉 THKey!=空后 rows={(rt4_filt['__WMSKeyTH__']!='').sum()}")

    # 退货输出需要 THKey 参与 join，但不匹配也要保留行
    rt4_filt["__RT4_ID__"] = range(len(rt4_filt))

    # 第一步：THKey -> 售后单号（left join 保留全部）
    merged1 = pd.merge(
        rt4_filt,
        rt3_df,
        how="left",
        left_on="__WMSKeyTH__",
        right_on=rt3_sa_after_col,
        suffixes=("_wms", "_rt3"),
    )

    matched1_mask = merged1[rt3_sa_after_col].notna() & (merged1[rt3_sa_after_col] != "")
    matched1 = merged1.loc[matched1_mask].copy()
    unmatched1 = merged1.loc[~matched1_mask].copy()
    if debug:
        st.write(f"RT_v2：step1 left join 匹配行数={len(matched1)}，未匹配行数={len(unmatched1)}")

    # 未匹配部分：再用 WMS 外部单号 -> 线上订单号/内部订单号
    # 为了避免重复字段冲突，unmatched部分只取 WMS 侧列
    wms_cols = list(rt4_filt.columns)
    wms_unmatched = unmatched1[wms_cols].copy()

    merged2_online = pd.DataFrame()
    merged3_internal = pd.DataFrame()
    remaining2 = wms_unmatched

    if rt3_online_col:
        tmp_online = pd.merge(
            wms_unmatched,
            rt3_df,
            how="left",
            left_on="__WMSExternalNorm__",
            right_on=rt3_online_col,
            suffixes=("_wms", "_rt3"),
        )
        matched_online_mask = tmp_online[rt3_online_col].notna() & (tmp_online[rt3_online_col] != "")
        merged2_online = tmp_online.loc[matched_online_mask].copy()
        remaining2 = tmp_online.loc[~matched_online_mask].copy()
        if debug:
            st.write(f"RT_v2：step2 线上订单号匹配行数={len(merged2_online)}，剩余未匹配={len(remaining2)}")

    if rt3_internal_col:
        # 对剩余未匹配再用内部订单号
        remaining2_wms = remaining2[wms_cols].copy() if set(wms_cols).issubset(set(remaining2.columns)) else remaining2.copy()
        tmp_internal = pd.merge(
            remaining2_wms,
            rt3_df,
            how="left",
            left_on="__WMSExternalNorm__",
            right_on=rt3_internal_col,
            suffixes=("_wms", "_rt3"),
        )
        matched_internal_mask = tmp_internal[rt3_internal_col].notna() & (tmp_internal[rt3_internal_col] != "")
        merged3_internal = tmp_internal.loc[matched_internal_mask].copy()
        remaining2 = tmp_internal.loc[~matched_internal_mask].copy()
        if debug:
            st.write(f"RT_v2：step2 内部订单号匹配行数={len(merged3_internal)}，最终未匹配={len(remaining2)}")

    # concat：保留所有行（含最终未匹配）
    frames = [matched1]
    if not merged2_online.empty:
        frames.append(merged2_online)
    if not merged3_internal.empty:
        frames.append(merged3_internal)
    # remaining2 现在要保证是 merged 形态（已经包含rt3列，或仅wms列均可）
    frames.append(remaining2)
    merged_rt = pd.concat(frames, ignore_index=True)

    # 去重防卫（避免 join 回填引入重复行）
    merged_rt = merged_rt.drop_duplicates(subset=["__RT4_ID__", wms_sku_col], keep="first")
    if debug:
        st.write(f"👉 两表 Merge 匹配后，成功合并的行数: {len(merged_rt)}")

    # 剔除商品关键词（如果列存在）
    if wms_sku_name_col and wms_sku_name_col in merged_rt.columns:
        merged_rt = merged_rt.loc[~_contains_keyword(merged_rt[wms_sku_name_col], SO_EXCLUDE_KEYWORDS)].copy()

    # 数值字段强制转换（NaN -> 0）
    qty_pos = pd.to_numeric(merged_rt[wms_qty_col], errors="coerce").fillna(0)
    amt_pos = pd.to_numeric(merged_rt[wms_amount_col], errors="coerce").fillna(0)

    # 退货表缺失 运费收入分摊 时当作 0：价税合计 = 金额 - 0
    actual_pay_pos = amt_pos

    # 退货输出前转负数：数量、金额都 * -1
    qty_out = -qty_pos
    pay_out = -actual_pay_pos

    # 构建基础输出
    shop_out = merged_rt[rt3_shop_col] if rt3_shop_col and rt3_shop_col in merged_rt.columns else "N/A"
    online_out = merged_rt[rt3_online_col] if rt3_online_col and rt3_online_col in merged_rt.columns else ""

    base_out = pd.DataFrame(
        {
            "店铺": shop_out,
            "商品编码": merged_rt[wms_sku_col],
            "实发数量": qty_out,
            "实际支付金额": pay_out,
            "线上订单号": online_out,
            "订单号": merged_rt.get(rt3_sa_after_col, ""),
        }
    )

    # 运费H101切分（如果WMS提供运费金额列）
    if wms_ship_fee_col and wms_ship_fee_col in merged_rt.columns:
        ship_fee_pos = pd.to_numeric(merged_rt[wms_ship_fee_col], errors="coerce").fillna(0)
        fee_mask = ship_fee_pos > 0
        if fee_mask.any():
            fee_rows = merged_rt.loc[fee_mask].copy()
            fee_shop = fee_rows[rt3_shop_col] if rt3_shop_col and rt3_shop_col in fee_rows.columns else "N/A"
            fee_online = fee_rows[rt3_online_col] if rt3_online_col and rt3_online_col in fee_rows.columns else ""
            fee_out = pd.DataFrame(
                {
                    "店铺": fee_shop,
                    "商品编码": ["H101"] * len(fee_rows),
                    "实发数量": [-1.0] * len(fee_rows),
                    "实际支付金额": (-ship_fee_pos.loc[fee_mask]).values,
                    "线上订单号": fee_online,
                    "订单号": fee_rows.get(rt3_sa_after_col, ""),
                }
            )
            base_out = pd.concat([base_out, fee_out], ignore_index=True)

    # 按负数后的实发数量拆分
    nonbao_df = base_out.loc[base_out["实发数量"] == 0].copy()
    main_df = base_out.loc[base_out["实发数量"] != 0].copy()

    # 输出前补齐 NaN：字符串为空，数值为0
    for col in ["店铺", "商品编码", "线上订单号", "订单号"]:
        if col in main_df.columns:
            main_df[col] = main_df[col].fillna("").astype(str)
        if col in nonbao_df.columns:
            nonbao_df[col] = nonbao_df[col].fillna("").astype(str)

    for col in ["实发数量", "实际支付金额"]:
        if col in main_df.columns:
            main_df[col] = pd.to_numeric(main_df[col], errors="coerce").fillna(0)
        if col in nonbao_df.columns:
            nonbao_df[col] = pd.to_numeric(nonbao_df[col], errors="coerce").fillna(0)

    return main_df, nonbao_df


def _compute_report_rows(
    df: pd.DataFrame,
    aux_map: Dict[str, Tuple[str, str]],
    is_return: bool,
) -> List[List[object]]:
    """
    返回给 openpyxl 的每行数据（不含表头），顺序严格为 25 列：
    [电商系统物料编码, 辅助自由项编码1, 辅助自由项编码2,
     主计量单位, 辅计量单位,
     数量, 税率, 含税单价, 无税单价, 税额, 无税金额, 价税合计,
     单品折扣, 整单折扣,
     客户订单号, 备注, 是否赠品, 扩展字段1~8, 发货仓库编码]
    """
    if df is None or df.empty:
        return []

    df = df.copy()

    # 强制类型
    if "实发数量" in df.columns:
        _ensure_numeric(df, "实发数量")
    else:
        df["实发数量"] = 0
    if "实际支付金额" in df.columns:
        _ensure_numeric(df, "实际支付金额")
    else:
        df["实际支付金额"] = 0

    # 为避免除零：税/单价字段先按0处理
    qty = pd.to_numeric(df["实发数量"], errors="coerce").fillna(0)
    price = pd.to_numeric(df["实际支付金额"], errors="coerce").fillna(0)

    # is_return 仅用于标识输出口径；正负号由上游数据逻辑控制
    # （SO为正，RT在 _enrich_rt_df 中已转换为负数）

    ex_tax = (price / 1.13).round(2)
    tax = (price - ex_tax).round(2)
    with np.errstate(divide="ignore", invalid="ignore"):
        # 单价精度要求：四舍五入到 2 位小数
        tax_incl_unit = np.where(qty != 0, (price / qty).round(2), 0.0)
        tax_excl_unit = np.where(qty != 0, (ex_tax / qty).round(2), 0.0)

    df["__qty__"] = qty
    df["__price__"] = price
    df["__ex_tax__"] = ex_tax
    df["__tax__"] = tax
    df["__tax_incl_unit__"] = tax_incl_unit
    df["__tax_excl_unit__"] = tax_excl_unit

    rows: List[List[object]] = []
    for i in range(len(df)):
        r = df.iloc[i]
        item_code = "" if pd.isna(r.get("商品编码", "")) else str(r.get("商品编码", "")).strip()
        aux_code1 = ""
        aux_code2 = ""
        main_code = item_code
        # 严格限制：必须是纯数字的长串，且包含 '00' 才允许切分
        if "00" in item_code and item_code.isdigit() and len(item_code) > 6:
            aux_code1 = item_code[6:]
            main_code = item_code[:6]

        # 使用完整编码查辅助表，查不到再默认 PCS
        main_u, sub_u = aux_map.get(item_code, ("PCS", "PCS"))
        if not main_u:
            main_u = "PCS"
        if not sub_u:
            sub_u = main_u

        quantity = float(r["__qty__"])
        amount_incl = float(r["__price__"])
        amount_excl = float(r["__ex_tax__"])
        amount_tax = float(r["__tax__"])
        unit_incl = float(r["__tax_incl_unit__"])
        unit_excl = float(r["__tax_excl_unit__"])
        client_order = _parse_first_order_id(r.get("线上订单号", ""))

        cols = [
            main_code,  # 电商系统物料编码
            aux_code1,  # 辅助自由项编码1
            aux_code2,  # 辅助自由项编码2
            main_u,  # 主计量单位
            sub_u,  # 辅计量单位
            quantity,  # 数量
            "13.00",  # 税率
            unit_incl,  # 含税单价
            unit_excl,  # 无税单价
            amount_tax,  # 税额
            amount_excl,  # 无税金额
            amount_incl,  # 价税合计
            "100%",  # 单品折扣
            "100%",  # 整单折扣
            client_order,  # 客户订单号
            "",  # 备注
            "否",  # 是否赠品
            "",  # 扩展字段1
            "",  # 扩展字段2
            "",  # 扩展字段3
            "",  # 扩展字段4
            "",  # 扩展字段5
            "",  # 扩展字段6
            "",  # 扩展字段7
            "2107",  # 发货仓库编码（扩展字段8）
        ]
        rows.append(cols)

        # 恢复：H101 运费单独成行（仅限发货单）
        freight_amt = float(r.get("运费金额", 0.0))
        if (not is_return) and freight_amt > 0:
            f_ex_tax = round(freight_amt / 1.13, 2)
            f_tax = round(freight_amt - f_ex_tax, 2)
            freight_cols = [
                "H101",  # 电商系统物料编码
                "",  # 辅助自由项编码1
                "",  # 辅助自由项编码2
                "PCS",  # 主计量单位
                "PCS",  # 辅计量单位
                1.0,  # 数量
                "13.00",  # 税率
                freight_amt,  # 含税单价（单件即总额）
                f_ex_tax,  # 无税单价
                f_tax,  # 税额
                f_ex_tax,  # 无税金额
                freight_amt,  # 价税合计
                "100%",  # 单品折扣
                "100%",  # 整单折扣
                client_order,  # 客户订单号
                "",  # 备注
                "否",  # 是否赠品
                "",  # 扩展字段1
                "",  # 扩展字段2
                "",  # 扩展字段3
                "",  # 扩展字段4
                "",  # 扩展字段5
                "",  # 扩展字段6
                "",  # 扩展字段7
                "2107",  # 发货仓库编码（扩展字段8）
            ]
            rows.append(freight_cols)

    return rows


def _dataframe_to_excel_bytes(
    df: pd.DataFrame,
    yesterday: date,
    aux_map: Dict[str, Tuple[str, str]],
    is_return: bool,
) -> BytesIO:
    if df is None:
        df = pd.DataFrame()
    df = df.copy()
    # 输出前统一 NaN 处理：字符串列为空，数值列为0
    for col in ["店铺", "商品编码", "线上订单号", "订单号"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)
    for col in ["实发数量", "实际支付金额"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    yesterday_slash = yesterday.strftime("%Y/%m/%d")
    yesterday_ymd = yesterday.strftime("%Y%m%d")

    output = BytesIO()
    wb: Workbook = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头：第1-2行
    ws["A1"] = "单据日期"
    ws["B1"] = "含税总金额"
    ws["C1"] = "交易类型"
    ws["D1"] = "流程类型"
    ws["E1"] = "收款协议"
    ws["F1"] = "部门编码"
    ws["G1"] = "客户编码"

    ws["A2"] = yesterday_slash
    ws["C2"] = "Cxx-TM02"
    ws["D2"] = "30-Cxx-04"
    ws["E2"] = "S003"
    ws["F2"] = "200101"
    ws["G2"] = "9701"

    # 第3行留空：默认即可

    # 表体：第4行起
    body_headers = [
        "电商系统物料编码",
        "辅助自由项编码1",
        "辅助自由项编码2",
        "主计量单位",
        "辅计量单位",
        "数量",
        "税率",
        "含税单价",
        "无税单价",
        "税额",
        "无税金额",
        "价税合计",
        "单品折扣",
        "整单折扣",
        "客户订单号",
        "备注",
        "是否赠品",
        "扩展字段1",
        "扩展字段2",
        "扩展字段3",
        "扩展字段4",
        "扩展字段5",
        "扩展字段6",
        "扩展字段7",
        "发货仓库编码",
    ]

    for col_idx, name in enumerate(body_headers, start=1):
        ws.cell(row=4, column=col_idx, value=name)

    report_rows = _compute_report_rows(df, aux_map=aux_map, is_return=is_return)
    data_start_row = 5
    for r_idx, row_data in enumerate(report_rows):
        excel_row = data_start_row + r_idx
        for c_idx, v in enumerate(row_data, start=1):
            ws.cell(row=excel_row, column=c_idx, value=v)

    # B2：价税合计列求和（列L）
    if report_rows:
        last_row = data_start_row + len(report_rows) - 1
        ws["B2"] = f"=SUM(L{data_start_row}:L{last_row})"
    else:
        ws["B2"] = 0

    # 简单列宽（不做复杂样式）
    col_widths = {
        1: 16,  # 电商系统物料编码
        2: 16,  # 辅助自由项编码1
        3: 16,  # 辅助自由项编码2
        4: 12,
        5: 12,
        6: 10,
        7: 8,
        8: 12,
        9: 12,
        10: 12,
        11: 12,
        12: 14,
        13: 10,
        14: 10,
        15: 18,
        16: 18,
        17: 10,
        18: 10,
        19: 10,
        20: 10,
        21: 10,
        22: 10,
        23: 10,
        24: 14,
        25: 14,
    }
    for col_idx in range(1, len(body_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 12)

    wb.save(output)
    output.seek(0)
    return output


def _dfs_to_concat(files: List[object], st_label: str) -> Optional[pd.DataFrame]:
    dfs = []
    for f in files or []:
        try:
            df = pd.read_excel(f)
            if df is None or df.empty:
                continue
            df = _normalize_columns(df)
            dfs.append(df)
        except Exception as e:
            st.warning(f"{st_label}读取失败：{e}")
    if not dfs:
        return None
    return pd.concat(dfs, ignore_index=True)


def _dfs_to_concat_by_name_keywords(
    files: List[object],
    st_label: str,
    name_keywords: Optional[List[str]],
) -> Optional[pd.DataFrame]:
    """
    按上传文件名宽容过滤读取（用于兜底识别文件）。
    若过滤后为空，会自动回退到“不过滤”。
    """
    if not files:
        return None

    name_keywords = name_keywords or []
    lowered_keywords = [k.lower() for k in name_keywords if k]

    def _match_one(f: object) -> bool:
        if not lowered_keywords:
            return True
        nm = str(getattr(f, "name", "") or "").lower()
        return any(k in nm for k in lowered_keywords)

    dfs: List[pd.DataFrame] = []
    for f in files or []:
        if not _match_one(f):
            continue
        try:
            if hasattr(f, "seek"):
                f.seek(0)
            df = pd.read_excel(f)
            if df is None or df.empty:
                continue
            df = _normalize_columns(df)
            dfs.append(df)
        except Exception as e:
            st.warning(f"{st_label}读取失败：{e}")

    if dfs:
        return pd.concat(dfs, ignore_index=True)

    # 兜底：名字过滤导致空，回退到读取所有
    dfs = []
    for f in files or []:
        try:
            if hasattr(f, "seek"):
                f.seek(0)
            df = pd.read_excel(f)
            if df is None or df.empty:
                continue
            df = _normalize_columns(df)
            dfs.append(df)
        except Exception as e:
            st.warning(f"{st_label}读取失败：{e}")

    if not dfs:
        return None
    return pd.concat(dfs, ignore_index=True)


def _zip_bytes(files: Dict[str, BytesIO]) -> BytesIO:
    zbuf = BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, bio in files.items():
            bio.seek(0)
            zf.writestr(name, bio.read())
    zbuf.seek(0)
    return zbuf


def _bucket_report_downloads(
    base_df: pd.DataFrame,
    aux_map: Dict[str, Tuple[str, str]],
    yesterday: date,
    is_return: bool,
) -> Dict[str, BytesIO]:
    """
    base_df: 统一列：店铺/商品编码/实发数量/实际支付金额/线上订单号/订单号
    """
    buckets = {
        "HAY-Tmall": [],
        "RED": [],
        "JD-STUDIO 9": [],
    }
    if base_df is not None and not base_df.empty and "店铺" in base_df.columns:
        for _, r in base_df.iterrows():
            bucket = _shop_bucket(r.get("店铺", ""))
            buckets[bucket].append(r)
    elif base_df is not None and not base_df.empty:
        # 无店铺列时，全部丢到兜底
        buckets["HAY-Tmall"] = [r for _, r in base_df.iterrows()]

    result: Dict[str, BytesIO] = {}
    for bucket_key, items in buckets.items():
        df_bucket = pd.DataFrame(items)
        if df_bucket.empty:
            df_bucket = pd.DataFrame(columns=["店铺", "商品编码", "实发数量", "实际支付金额", "线上订单号", "订单号"])
        file_bytes = _dataframe_to_excel_bytes(df_bucket, yesterday=yesterday, aux_map=aux_map, is_return=is_return)
        result[bucket_key] = file_bytes
    return result


def main() -> None:
    st.set_page_config(page_title="电商财务对账与报表生成", layout="wide")
    st.title("电商财务对账与报表生成")
    st.caption("支持上传多个 .xlsx/.xls 文件，内存处理并直接下载 SO/RT 报表（含 ZIP 打包）。")

    with st.expander("1) 上传文件（允许为空）", expanded=True):
        so1_files = st.file_uploader(
            "销售底表1（聚水潭出库表）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="so1_files",
        )
        so2_files = st.file_uploader(
            "销售底表2（WMS发货明细）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="so2_files",
        )
        rt3_files = st.file_uploader(
            "退货底表3（聚水潭退货表）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="rt3_files",
        )
        rt4_files = st.file_uploader(
            "退货底表4（WMS收货明细）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="rt4_files",
        )
        aux_files = st.file_uploader(
            "辅助表（匹配物料单位）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="aux_files",
        )
        manual_files = st.file_uploader(
            "手工单表（独立业务）",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="manual_files",
        )

    st.divider()

    colA, colB = st.columns([1, 2])
    with colA:
        run = st.button("生成报表", type="primary")
    with colB:
        st.info("提示：你上传多少就处理多少；缺失的部分会跳过并继续生成可下载结果。")

    if not run:
        st.stop()

    yesterday = date.today() - timedelta(days=1)
    yesterday_prefix = yesterday.strftime("%Y%m%d")

    debug = st.checkbox("显示调试信息（用于定位为什么为0）", value=False)

    with st.spinner("正在读取/处理数据，请稍候..."):
        aux_df = _dfs_to_concat(aux_files, "辅助表")
        aux_map = _build_aux_map(aux_df)

        st.write(f"辅助表：{0 if aux_df is None else len(aux_df)} 行")

        # 宽容文件名识别（如果文件名不规范会自动回退为“不过滤”，避免整批空数据）
        so1_df = (
            _dfs_to_concat_by_name_keywords(
                so1_files,
                "销售底表1",
                ["销售出库", "底表1", "出库表", "聚水潭出库"],
            )
            if so1_files
            else None
        )
        so2_df = (
            _dfs_to_concat_by_name_keywords(
                so2_files,
                "销售底表2",
                ["发货明细", "发货", "底表2", "WMS发货"],
            )
            if so2_files
            else None
        )
        rt3_df = (
            _dfs_to_concat_by_name_keywords(
                rt3_files,
                "退货底表3",
                ["退货", "底表3", "售后", "聚水潭退货"],
            )
            if rt3_files
            else None
        )
        rt4_df = (
            _dfs_to_concat_by_name_keywords(
                rt4_files,
                "退货底表4",
                ["收货明细", "收货", "底表4", "WMS收货"],
            )
            if rt4_files
            else None
        )
        # RT字段匹配对“列名是否带不可见空格”极其敏感：按要求立即清理表头空格
        if rt3_df is not None:
            rt3_df.columns = rt3_df.columns.astype(str).str.strip()
        if rt4_df is not None:
            rt4_df.columns = rt4_df.columns.astype(str).str.strip()
        manual_df = (
            _dfs_to_concat_by_name_keywords(
                manual_files,
                "手工单表",
                ["手工", "手工单"],
            )
            if manual_files
            else None
        )

        if debug:
            st.write("=== Debug：上传文件读取结果 ===")
            st.write(f"aux_df：{None if aux_df is None else aux_df.shape}, sample_cols={[] if aux_df is None else list(aux_df.columns[:20])}")
            st.write(f"so1_df：{None if so1_df is None else so1_df.shape}, sample_cols={[] if so1_df is None else list(so1_df.columns[:20])}")
            st.write(f"so2_df：{None if so2_df is None else so2_df.shape}, sample_cols={[] if so2_df is None else list(so2_df.columns[:20])}")
            st.write(f"rt3_df：{None if rt3_df is None else rt3_df.shape}, sample_cols={[] if rt3_df is None else list(rt3_df.columns[:20])}")
            st.write(f"rt4_df：{None if rt4_df is None else rt4_df.shape}, sample_cols={[] if rt4_df is None else list(rt4_df.columns[:20])}")
            st.write(f"manual_df：{None if manual_df is None else manual_df.shape}, sample_cols={[] if manual_df is None else list(manual_df.columns[:20])}")

        so_base_df = pd.DataFrame()
        if so1_df is not None and so2_df is not None:
            so_base_df = _enrich_so_df(so1_df, so2_df, aux_map=aux_map, debug=debug)
        st.write(f"SO匹配后行数：{len(so_base_df)}")

        rt_main_df = pd.DataFrame()
        rt_nonbao_df = pd.DataFrame()
        if rt3_df is not None and rt4_df is not None:
            rt_main_df, rt_nonbao_df = _enrich_rt_df_v2(rt3_df, rt4_df, aux_map=aux_map, debug=debug)
        st.write(f"RT主退货行数：{len(rt_main_df) if rt_main_df is not None else 0}，RT非保行数：{len(rt_nonbao_df) if rt_nonbao_df is not None else 0}")

        # 手工单：尽量按“同格式处理（SO正 RT负）”
        manual_so_df = pd.DataFrame()
        manual_rt_df = pd.DataFrame()
        if manual_df is not None and not manual_df.empty:
            m = _normalize_columns(manual_df)
            qty_col = _first_existing_col(m, ["实发数量", "数量"])
            amt_col = _first_existing_col(m, ["实发金额", "金额"])
            ship_income_col = _first_existing_col(m, ["运费收入分摊", "运费分摊"])
            sku_name_col = _first_existing_col(m, ["商品简称", "商品名称", "商品描述"])
            sku_code_col = _first_existing_col(m, ["商品编码", "物料编码", "SKU编码", "电商系统物料编码"])
            shop_col = _first_existing_col(m, ["店铺", "店铺名称", "平台店铺"])
            online_order_col = _first_existing_col(m, ["线上订单号", "客户订单号", "线上订单", "订单号线上", "线上订单编号"])
            trx_col = _first_existing_col(m, ["交易类型", "单据类型", "业务类型", "类型", "收发标识"])

            if qty_col and amt_col and sku_code_col:
                _ensure_numeric(m, qty_col)
                _ensure_numeric(m, amt_col)
                if ship_income_col:
                    _ensure_numeric(m, ship_income_col)
                    m["实际支付金额"] = m[amt_col] - m[ship_income_col]
                else:
                    m["实际支付金额"] = m[amt_col]
                    m["运费收入分摊"] = 0

                # 商品排除
                if sku_name_col:
                    m = m.loc[~_contains_keyword(m[sku_name_col], SO_EXCLUDE_KEYWORDS)].copy()

                m["商品编码"] = m[sku_code_col]
                m["实发数量"] = m[qty_col]
                m["线上订单号"] = m[online_order_col].apply(_parse_first_order_id) if online_order_col else ""
                m["店铺"] = m[shop_col] if shop_col else "N/A"
                m["订单号"] = (
                    m.get("出仓单号", None)
                    if "出仓单号" in m.columns
                    else m.get("售后单号", None)
                    if "售后单号" in m.columns
                    else ""
                )

                # 发/退识别：优先看交易类型字段
                if trx_col:
                    trx_series = m[trx_col].astype(str)
                    manual_rt_df = m.loc[trx_series.str.contains("退", na=False)].copy()
                    manual_so_df = m.loc[~trx_series.str.contains("退", na=False)].copy()
                else:
                    # 兜底：用实发数量正负（如果为负则当作退）
                    manual_rt_df = m.loc[m["实发数量"] < 0].copy()
                    manual_so_df = m.loc[m["实发数量"] >= 0].copy()

                # 按需求：发货为正，退货为负
                if not manual_so_df.empty:
                    manual_so_df["实发数量"] = pd.to_numeric(manual_so_df["实发数量"], errors="coerce").fillna(0).abs()
                    manual_so_df["实际支付金额"] = pd.to_numeric(manual_so_df["实际支付金额"], errors="coerce").fillna(0).abs()
                if not manual_rt_df.empty:
                    # 退货按要求转为负数：数量/金额都 * -1
                    manual_rt_df["实发数量"] = -pd.to_numeric(manual_rt_df["实发数量"], errors="coerce").fillna(0).abs()
                    manual_rt_df["实际支付金额"] = -pd.to_numeric(manual_rt_df["实际支付金额"], errors="coerce").fillna(0).abs()

                manual_so_df = manual_so_df[["店铺", "商品编码", "实发数量", "实际支付金额", "线上订单号", "订单号"]].copy()
                manual_rt_df = manual_rt_df[["店铺", "商品编码", "实发数量", "实际支付金额", "线上订单号", "订单号"]].copy()

        # 组装输出文件
        out_files: Dict[str, BytesIO] = {}

        # SO分流
        so_buckets = _bucket_report_downloads(
            base_df=so_base_df if so_base_df is not None else pd.DataFrame(),
            aux_map=aux_map,
            yesterday=yesterday,
            is_return=False,
        )
        # 兜底输出：仍需so-HAY-Tmall/so-RED/so-JD-STUDIO 9（即使没有数据）
        so_hay_tmall = so_buckets.get("HAY-Tmall")
        so_red = so_buckets.get("RED")
        so_jd = so_buckets.get("JD-STUDIO 9")
        if so_hay_tmall:
            out_files[f"{yesterday_prefix}-so-HAY-Tmall.xlsx"] = so_hay_tmall
        else:
            out_files[f"{yesterday_prefix}-so-HAY-Tmall.xlsx"] = _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=False)
        if so_red:
            out_files[f"{yesterday_prefix}-so-RED.xlsx"] = so_red
        else:
            out_files[f"{yesterday_prefix}-so-RED.xlsx"] = _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=False)
        if so_jd:
            out_files[f"{yesterday_prefix}-so-JD-STUDIO 9.xlsx"] = so_jd
        else:
            out_files[f"{yesterday_prefix}-so-JD-STUDIO 9.xlsx"] = _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=False)

        # so-手工单.xlsx：手工单不再分店铺，按表内SO交易类型汇总
        if manual_so_df is not None and not manual_so_df.empty:
            out_files[f"{yesterday_prefix}-so-手工单.xlsx"] = _dataframe_to_excel_bytes(manual_so_df, yesterday, aux_map, is_return=False)
        else:
            out_files[f"{yesterday_prefix}-so-手工单.xlsx"] = _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=False)

        # RT分流（主退货）
        rt_main_buckets = _bucket_report_downloads(
            base_df=rt_main_df if rt_main_df is not None else pd.DataFrame(),
            aux_map=aux_map,
            yesterday=yesterday,
            is_return=True,
        )
        rt_hay_tmall = rt_main_buckets.get("HAY-Tmall")
        rt_red = rt_main_buckets.get("RED")
        rt_jd = rt_main_buckets.get("JD-STUDIO 9")
        out_files[f"{yesterday_prefix}-rt-HAY-Tmall.xlsx"] = (
            rt_hay_tmall if rt_hay_tmall else _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=True)
        )
        out_files[f"{yesterday_prefix}-rt-RED.xlsx"] = (
            rt_red if rt_red else _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=True)
        )
        out_files[f"{yesterday_prefix}-rt-JD-STUDIO 9.xlsx"] = (
            rt_jd if rt_jd else _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=True)
        )

        # rt-非保.xlsx：仅实发数量==0的退货单
        out_files[f"{yesterday_prefix}-rt-非保.xlsx"] = _dataframe_to_excel_bytes(
            rt_nonbao_df if rt_nonbao_df is not None else pd.DataFrame(), yesterday, aux_map, is_return=True
        )

        # rt-手工单.xlsx
        if manual_rt_df is not None and not manual_rt_df.empty:
            out_files[f"{yesterday_prefix}-rt-手工单.xlsx"] = _dataframe_to_excel_bytes(manual_rt_df, yesterday, aux_map, is_return=True)
        else:
            out_files[f"{yesterday_prefix}-rt-手工单.xlsx"] = _dataframe_to_excel_bytes(pd.DataFrame(), yesterday, aux_map, is_return=True)

        # ZIP打包
        zip_buf = _zip_bytes(out_files)

    st.success("报表已生成，可以开始下载。")
    st.write(
        {
            "so匹配行数": int(len(so_base_df)) if so_base_df is not None else 0,
            "rt主退货行数": int(len(rt_main_df)) if rt_main_df is not None else 0,
            "rt非保退货行数": int(len(rt_nonbao_df)) if rt_nonbao_df is not None else 0,
            "手工单行数": int(len(manual_df)) if manual_df is not None else 0,
        }
    )

    st.subheader("打包下载（ZIP）")
    st.download_button(
        label=f"{yesterday_prefix}-报表ZIP下载",
        data=zip_buf.getvalue(),
        file_name=f"{yesterday_prefix}-报表.zip",
        mime="application/zip",
        key="download_zip",
    )

    st.subheader("单文件下载")
    # 为了界面整洁：两列展示下载按钮
    download_items = list(out_files.items())
    cols = st.columns(2)
    for idx, (fname, bio) in enumerate(download_items):
        col = cols[idx % 2]
        bio.seek(0)
        col.download_button(
            label=fname,
            data=bio.getvalue(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{fname}",
        )


if __name__ == "__main__":
    main()

